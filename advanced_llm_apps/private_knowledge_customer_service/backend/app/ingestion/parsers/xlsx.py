from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.parsers.base import ParsedSection


class XlsxParser:
    def parse(self, path: Path) -> list[ParsedSection]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sections: list[ParsedSection] = []
        try:
            for sheet in workbook.worksheets:
                rows = [
                    "\t".join("" if value is None else str(value) for value in row).rstrip()
                    for row in sheet.iter_rows(values_only=True)
                ]
                populated_rows = [row for row in rows if row.strip()]
                if not populated_rows:
                    continue
                sections.append(
                    ParsedSection(
                        text="\n".join(populated_rows),
                        locator={
                            "sheet": sheet.title,
                            "row_start": 1,
                            "row_end": len(rows),
                        },
                    )
                )
        finally:
            workbook.close()
        return sections

